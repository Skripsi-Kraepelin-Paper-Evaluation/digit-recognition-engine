from flask import Blueprint, jsonify, request
import os
import json
import shutil
import tempfile
from openpyxl import load_workbook
from models import predicted_digit_answer as answer

class ImportHandler:
    def __init__(self, persistent_path='./persistent', excel_template_path='./asset/template.template.xlsx',questions_fixed=[]):
        self.persistent_path = persistent_path
        self.excel_template_path = excel_template_path
        self.questions_fixed = questions_fixed

    def handle_import(self, uploaded_file):
        temp_file_path = None
        try:
            # Get OS temp directory
            temp_dir = tempfile.gettempdir()
            
            # Save uploaded file to temp directory
            filename = uploaded_file.filename
            temp_file_path = os.path.join(temp_dir, filename)
            uploaded_file.save(temp_file_path)
            
            print(f"Saved uploaded file to: {temp_file_path}")
            
            # Load excel
            wb = load_workbook(temp_file_path, data_only=True)
            sheet = wb["Master"]
            
            # Validate payload and scan columns
            if sheet is not None:
                try:
                    answers = self.import_answers_from_sheet(sheet)
                    
                    # Clean up temp file
                    if temp_file_path and os.path.exists(temp_file_path):
                        os.remove(temp_file_path)
                        print(f"Cleaned up temporary file: {temp_file_path}")
                    
                    return {
                        "success": True,
                        'questions': self.questions_fixed,
                        "answers": [a.serialize_obj() for a in answers],
                        'total_questions': len(self.questions_fixed),
                        "total_answers": len(answers)
                    }
                except Exception as e:
                    print(f"Error during scanning: {e}")
                    # Clean up on error
                    if temp_file_path and os.path.exists(temp_file_path):
                        os.remove(temp_file_path)
                    raise ValueError(f"Invalid Excel: {str(e)}")
            else:
                if temp_file_path and os.path.exists(temp_file_path):
                    os.remove(temp_file_path)
                raise ValueError("Invalid Excel: Master sheet not found")
                
        except Exception as e:
            # Ensure cleanup on any error
            if temp_file_path and os.path.exists(temp_file_path):
                try:
                    os.remove(temp_file_path)
                except:
                    pass
            raise e

    def import_answers_from_sheet(self, sheet):
        """
        Scan specific columns from row 3 to 56, bottom-up, left-to-right.
        Columns: D, H, L, P, T, X, AB, AF, AJ, AN, AR, AV, AZ, BD, BH, BL, BP, BT, BX, CB, CF, CJ, CN, CR, CV, CZ, DD, DH, DL, DP, DT, DX, EB, EF, EJ, EN, ER, EV, EZ, FD, FH, FL, FP, FT, FX
        
        Args:
            sheet: openpyxl worksheet object
        
        Returns:
            list: 3D array where each element is [column_index, row, value]
        """
        columns = []
        
        # Generate column letters: every 4th column starting from D
        start_col = 4  # Column D
        for i in range(40):  # only take 40 columns (D to FX with step of 4)
            col_num = start_col + (i * 4)
            col_letter = self.get_column_letter(col_num)
            columns.append(col_letter)
        
        print(f"Scanning columns: {columns}")
        
        result_answers = []
        
        # Scan left to right
        for col_idx, col_letter in enumerate(columns):
            column_data = []
            totalRow = 57
            upperRow = 2
            
            # Scan bottom-up (from row 56 to row 7)
            for row in range(totalRow, upperRow, -1):
                cell = sheet[f'{col_letter}{row}']
                cell_value = cell.value
                
                # Process the cell value
                if cell_value is None or cell_value == '':
                    processed_value = 'N/A'
                elif isinstance(cell_value, (int, float)):
                    processed_value = int(cell_value)
                elif isinstance(cell_value, str):
                    cell_value_upper = cell_value.strip().upper()
                    if cell_value_upper in ['SKIPPED', 'N/A']:
                        processed_value = cell_value_upper
                    else:
                        # Try to convert to integer
                        try:
                            processed_value = int(cell_value)
                        except ValueError:
                            processed_value = 'N/A'
                else:
                    processed_value = 'N/A'

                if processed_value == 'N/A':
                    processed_value = 'BLANK'

                valInt = processed_value
                if processed_value == 'BLANK':
                    valInt = -1
                
                # Append to column data: [column_index, value]
                if processed_value != 'SKIPPED':
                    model_class = answer.PredictedDigitAnswer
                    result_answers.append(model_class(
                        digit=valInt,
                        accuracy=1.0,
                        column=col_idx,
                        row=totalRow-row,
                        need_manual_check=False,
                        checked=True,
                        is_blank=processed_value == 'BLANK'
                    ))
            # Append entire column to result
            print(f"Scanned column {col_letter} ({col_idx}): {len(column_data)} cells")
        
        return result_answers
    
    def get_column_letter(self, col_num):
        """Convert column number to Excel column letter (1=A, 2=B, etc.)"""
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(col_num % 26 + 65) + result
            col_num //= 26
        return result


def create_import_blueprint(cfg):
    import_handler = ImportHandler(persistent_path=cfg.persistent_path,questions_fixed=cfg.questions_fixed)
    import_bp = Blueprint('import_controller', __name__)

    @import_bp.route('/import/<filename>', methods=['POST'])
    def importer(filename):
        try:
            if 'file' not in request.files:
                return jsonify({"error": "No file provided"}), 400
            
            uploaded_file = request.files['file']
            if uploaded_file.filename == '':
                return jsonify({"error": "No file selected"}), 400

            result = import_handler.handle_import(uploaded_file)
            return jsonify(result), 200
        except ValueError as ve:
            return jsonify({"error": str(ve)}), 400
        except Exception as e:
            return jsonify({"error": f"Processing failed: {str(e)}"}), 500
    
    return import_bp