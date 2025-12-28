from flask import Blueprint, jsonify, request, send_file
import os
import json
import shutil
import tempfile
from openpyxl import load_workbook
from io import BytesIO

class ExportHandler:
    def __init__(self, persistent_path='./persistent', template_path='./asset/template.xlsx'):
        self.persistent_path = persistent_path
        self.template_path = template_path

    def handle_export(self, filename):
        try:
            # parsing payload
            payload = request.get_json()
            if not payload:
                raise ValueError("No JSON payload received")
            
            filename = payload.get('filename')
            if not filename:
                raise ValueError("Missing 'filename' in request")
                
            question_arrays = payload.get('questionArrays')
            if not question_arrays:
                raise ValueError("Missing 'questionArrays' in request")
                
            answer_arrays = payload.get('answerArrays')
            if not answer_arrays:
                raise ValueError("Missing 'answerArrays' in request")

            # Copy template to temp file
            temp_dir = tempfile.gettempdir()
            temp_file_path = os.path.join(temp_dir, f"{filename}_export.xlsx")
            shutil.copy2(self.template_path, temp_file_path)
            
            # Load and edit template
            wb = load_workbook(temp_file_path)
            sheet = wb["Master"]
            
            # Map answers to excel
            self.map_answers_to_sheet(sheet, answer_arrays)
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Clean up temp file
            if os.path.exists(temp_file_path):
                os.remove(temp_file_path)
            
            return output
            
        except Exception as e:
            raise Exception(f"Error processing request: {str(e)}")
    
    def map_answers_to_sheet(self, sheet, answer_arrays):
        """Map answer arrays to Excel sheet using same column mapping as import"""
        # Generate column letters: every 4th column starting from D
        columns = []
        start_col = 4  # Column D
        for i in range(40):  # 40 columns (D to FX with step of 4)
            col_num = start_col + (i * 4)
            col_letter = self.get_column_letter(col_num)
            columns.append(col_letter)
        
        # Map answers to cells
        for col_idx, answer_col in enumerate(answer_arrays):
            if col_idx >= len(columns):
                break
                
            col_letter = columns[col_idx]
            totalRow = 57
            upperRow = 2
            
            # Write answers bottom-up (reverse order)
            for row_idx, answer in enumerate(answer_col):
                excel_row = totalRow - row_idx
                if excel_row <= upperRow:
                    break
                    
                cell = sheet[f'{col_letter}{excel_row}']
                
                # Convert answer to appropriate value
                if answer == 'BLANK' or answer == '' or answer is None:
                    cell.value = 'N/A'
                elif answer == 'SKIPPED':
                    cell.value = 'SKIPPED'
                else:
                    try:
                        cell.value = int(answer)
                    except (ValueError, TypeError):
                        cell.value = None
    
    def get_column_letter(self, col_num):
        """Convert column number to Excel column letter (1=A, 2=B, etc.)"""
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(col_num % 26 + 65) + result
            col_num //= 26
        return result



def create_export_blueprint(cfg):
    export_handler = ExportHandler(persistent_path=cfg.persistent_path)
    export_bp = Blueprint('export_controller', __name__)

    @export_bp.route('/export/<filename>', methods=['POST'])
    def export(filename):
        try:
            output = export_handler.handle_export(filename)
            return send_file(
                output,
                as_attachment=True,
                download_name=f"{filename}_kraepelin_answers.xlsx",
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        except Exception as e:
            return jsonify({"error": f"Processing failed: {str(e)}"}), 500
    return export_bp
    